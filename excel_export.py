import os
import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import (
    LineChart, BarChart, ScatterChart,
    Reference, Series
)
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.series import DataPoint
from openpyxl.worksheet.datavalidation import DataValidation
from datetime import datetime
from typing import Dict, List, Optional, Any
from dataclasses import dataclass

@dataclass
class ExperimentRecord:
    experiment_id: str
    experiment_name: str
    data_type: str
    sensor_name: str
    threshold_parameter: float
    raw_metric: float
    normalized_metric: float
    passed: bool
    confidence: float
    adjustment_reason: str
    timestamp: str
    additional_info: Dict[str, Any] = None

@dataclass
class ComparisonResult:
    experiment_id: str
    method_name: str
    noise_type: str
    rmse_distance: float
    rmse_angle: float
    rmse_velocity: float
    rmse_emissivity: float
    rmse_reflectivity: float
    classification_acc: float
    f1_score: float
    inference_time_ms: float
    memory_mb: float
    overall_score: float
    timestamp: str

class ExcelExporter:
    def __init__(self, output_dir: str = "./experiment_results"):
        self.output_dir = output_dir
        self._ensure_directory(output_dir)
    
    def _ensure_directory(self, path: str):
        if not os.path.exists(path):
            os.makedirs(path)
    
    def _create_header_style(self):
        header_font = Font(name='Calibri', bold=True, size=11, color='FFFFFF')
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        return {
            'font': header_font,
            'fill': header_fill,
            'alignment': header_alignment,
            'border': thin_border
        }
    
    def _create_cell_style(self, passed: bool = True):
        cell_font = Font(name='Calibri', size=10)
        cell_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        if passed:
            cell_fill = PatternFill(start_color='D5E8D4', end_color='D5E8D4', fill_type='solid')
        else:
            cell_fill = PatternFill(start_color='F8CECC', end_color='F8CECC', fill_type='solid')
        
        return {
            'font': cell_font,
            'fill': cell_fill,
            'alignment': cell_alignment,
            'border': thin_border
        }
    
    def _create_dataframe_style(self):
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        return {
            'border': thin_border,
            'alignment': Alignment(horizontal='center', vertical='center')
        }
    
    def export_validation_results(self, records: List[ExperimentRecord], 
                                 filename: str = None) -> str:
        if filename is None:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"validation_results_{timestamp}.xlsx"
        
        filepath = os.path.join(self.output_dir, filename)
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Validation Results"
        
        header_style = self._create_header_style()
        cell_style_pass = self._create_cell_style(passed=True)
        cell_style_fail = self._create_cell_style(passed=False)
        
        headers = [
            'Experiment ID', 'Experiment Name', 'Data Type', 'Sensor Name',
            'Threshold Parameter', 'Raw Metric', 'Normalized Metric',
            'Passed', 'Confidence (%)', 'Adjustment Reason', 'Timestamp'
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            for key, value in header_style.items():
                setattr(cell, key, value)
        
        row = 2
        for record in records:
            ws.cell(row=row, column=1, value=record.experiment_id)
            ws.cell(row=row, column=2, value=record.experiment_name)
            ws.cell(row=row, column=3, value=record.data_type)
            ws.cell(row=row, column=4, value=record.sensor_name)
            ws.cell(row=row, column=5, value=record.threshold_parameter)
            ws.cell(row=row, column=6, value=record.raw_metric)
            ws.cell(row=row, column=7, value=record.normalized_metric)
            
            passed_cell = ws.cell(row=row, column=8, value="YES" if record.passed else "NO")
            style = cell_style_pass if record.passed else cell_style_fail
            for key, value in style.items():
                setattr(passed_cell, key, value)
            
            ws.cell(row=row, column=9, value=record.confidence * 100)
            ws.cell(row=row, column=10, value=record.adjustment_reason)
            ws.cell(row=row, column=11, value=record.timestamp)
            
            if record.additional_info:
                info_str = "; ".join(f"{k}: {v}" for k, v in record.additional_info.items())
                ws.cell(row=row, column=12, value=info_str)
            
            row += 1
        
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[chr(64 + col)].width = 15
        
        ws.column_dimensions['J'].width = 30
        
        ws.auto_filter.ref = ws.dimensions
        
        self._add_validation_summary_sheet(wb, records)
        
        self._add_validation_charts(wb, records)
        
        wb.save(filepath)
        
        return filepath
    
    def _add_validation_summary_sheet(self, wb: Workbook, records: List[ExperimentRecord]):
        ws = wb.create_sheet(title="Validation Summary")
        
        header_style = self._create_header_style()
        
        summary_data = []
        for sensor in set(r.sensor_name for r in records):
            sensor_records = [r for r in records if r.sensor_name == sensor]
            pass_count = sum(1 for r in sensor_records if r.passed)
            total = len(sensor_records)
            pass_rate = pass_count / total * 100 if total > 0 else 0
            avg_confidence = np.mean([r.confidence for r in sensor_records]) * 100
            
            summary_data.append({
                'Sensor Name': sensor,
                'Passed': pass_count,
                'Total': total,
                'Pass Rate (%)': pass_rate,
                'Avg Confidence (%)': avg_confidence
            })
        
        df = pd.DataFrame(summary_data)
        
        headers = list(df.columns)
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            for key, value in header_style.items():
                setattr(cell, key, value)
        
        for r, row_data in enumerate(df.itertuples(index=False), 2):
            for c, value in enumerate(row_data, 1):
                ws.cell(row=r, column=c, value=value)
        
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[chr(64 + col)].width = 18
        
        ws.auto_filter.ref = ws.dimensions
    
    def _add_validation_charts(self, wb: Workbook, records: List[ExperimentRecord]):
        ws = wb.create_sheet(title="Validation Charts")
        
        df = pd.DataFrame([{
            'Sensor': r.sensor_name,
            'Normalized Metric': r.normalized_metric,
            'Threshold': r.threshold_parameter,
            'Passed': 1 if r.passed else 0
        } for r in records])
        
        data_start_row = 2
        headers = list(df.columns)
        
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
        
        for r, row_data in enumerate(df.itertuples(index=False), data_start_row):
            for c, value in enumerate(row_data, 1):
                ws.cell(row=r, column=c, value=value)
        
        bar_chart = BarChart()
        bar_chart.title = "Normalized Metric vs Threshold"
        bar_chart.x_axis.title = "Sensor"
        bar_chart.y_axis.title = "Value"
        bar_chart.style = 10
        
        data = Reference(ws, min_col=2, min_row=data_start_row - 1, max_row=data_start_row - 1 + len(df))
        cats = Reference(ws, min_col=1, min_row=data_start_row, max_row=data_start_row - 1 + len(df))
        
        bar_chart.add_data(data, titles_from_data=True)
        bar_chart.set_categories(cats)
        
        bar_chart2 = BarChart()
        data2 = Reference(ws, min_col=3, min_row=data_start_row - 1, max_row=data_start_row - 1 + len(df))
        bar_chart2.add_data(data2, titles_from_data=True)
        
        bar_chart += bar_chart2
        
        ws.add_chart(bar_chart, "F2")
        
        scatter_chart = ScatterChart()
        scatter_chart.title = "Metric vs Confidence"
        scatter_chart.x_axis.title = "Normalized Metric"
        scatter_chart.y_axis.title = "Confidence"
        scatter_chart.style = 12
        
        x_values = Reference(ws, min_col=2, min_row=data_start_row, max_row=data_start_row - 1 + len(df))
        y_values = Reference(ws, min_col=4, min_row=data_start_row, max_row=data_start_row - 1 + len(df))
        
        series = Series(y_values, x_values, title="Confidence")
        scatter_chart.series.append(series)
        
        ws.add_chart(scatter_chart, "F25")
    
    def export_comparison_results(self, results: List[ComparisonResult],
                                  filename: str = None) -> str:
        if filename is None:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"comparison_results_{timestamp}.xlsx"
        
        filepath = os.path.join(self.output_dir, filename)
        
        wb = Workbook()
        
        self._add_filtering_sheet(wb, results)
        self._add_inversion_sheet(wb, results)
        self._add_efficiency_sheet(wb, results)
        self._add_overall_summary_sheet(wb, results)
        
        wb.save(filepath)
        
        return filepath
    
    def _add_filtering_sheet(self, wb: Workbook, results: List[ComparisonResult]):
        ws = wb.create_sheet(title="Filtering Performance")
        
        header_style = self._create_header_style()
        
        headers = ['Experiment ID', 'Method', 'Noise Type', 
                   'RMSE Distance', 'RMSE Angle', 'RMSE Velocity', 'Timestamp']
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            for key, value in header_style.items():
                setattr(cell, key, value)
        
        row = 2
        for result in results:
            ws.cell(row=row, column=1, value=result.experiment_id)
            ws.cell(row=row, column=2, value=result.method_name)
            ws.cell(row=row, column=3, value=result.noise_type)
            ws.cell(row=row, column=4, value=result.rmse_distance)
            ws.cell(row=row, column=5, value=result.rmse_angle)
            ws.cell(row=row, column=6, value=result.rmse_velocity)
            ws.cell(row=row, column=7, value=result.timestamp)
            row += 1
        
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[chr(64 + col)].width = 15
        
        ws.auto_filter.ref = ws.dimensions
        
        bar_chart = BarChart()
        bar_chart.title = "Filtering RMSE Comparison"
        bar_chart.x_axis.title = "Method"
        bar_chart.y_axis.title = "RMSE"
        bar_chart.style = 10
        
        data_range = Reference(ws, min_col=4, min_row=1, max_row=row - 1)
        cats = Reference(ws, min_col=2, min_row=2, max_row=row - 1)
        
        bar_chart.add_data(data_range, titles_from_data=True)
        bar_chart.set_categories(cats)
        
        ws.add_chart(bar_chart, "I2")
    
    def _add_inversion_sheet(self, wb: Workbook, results: List[ComparisonResult]):
        ws = wb.create_sheet(title="Inversion Performance")
        
        header_style = self._create_header_style()
        
        headers = ['Experiment ID', 'Method', 'Noise Type',
                   'Emissivity RMSE', 'Reflectivity RMSE',
                   'Classification Acc (%)', 'F1 Score', 'Timestamp']
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            for key, value in header_style.items():
                setattr(cell, key, value)
        
        row = 2
        for result in results:
            ws.cell(row=row, column=1, value=result.experiment_id)
            ws.cell(row=row, column=2, value=result.method_name)
            ws.cell(row=row, column=3, value=result.noise_type)
            ws.cell(row=row, column=4, value=result.rmse_emissivity)
            ws.cell(row=row, column=5, value=result.rmse_reflectivity)
            ws.cell(row=row, column=6, value=result.classification_acc)
            ws.cell(row=row, column=7, value=result.f1_score)
            ws.cell(row=row, column=8, value=result.timestamp)
            row += 1
        
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[chr(64 + col)].width = 18
        
        ws.auto_filter.ref = ws.dimensions
        
        bar_chart = BarChart()
        bar_chart.title = "Inversion Accuracy Comparison"
        bar_chart.x_axis.title = "Method"
        bar_chart.y_axis.title = "RMSE"
        bar_chart.style = 11
        
        data_range = Reference(ws, min_col=4, min_row=1, max_row=row - 1)
        cats = Reference(ws, min_col=2, min_row=2, max_row=row - 1)
        
        bar_chart.add_data(data_range, titles_from_data=True)
        bar_chart.set_categories(cats)
        
        bar_chart2 = BarChart()
        data_range2 = Reference(ws, min_col=5, min_row=1, max_row=row - 1)
        bar_chart2.add_data(data_range2, titles_from_data=True)
        
        bar_chart += bar_chart2
        
        ws.add_chart(bar_chart, "J2")
    
    def _add_efficiency_sheet(self, wb: Workbook, results: List[ComparisonResult]):
        ws = wb.create_sheet(title="Efficiency")
        
        header_style = self._create_header_style()
        
        headers = ['Experiment ID', 'Method', 'Noise Type',
                   'Inference Time (ms)', 'Memory (MB)', 'Overall Score', 'Timestamp']
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            for key, value in header_style.items():
                setattr(cell, key, value)
        
        row = 2
        for result in results:
            ws.cell(row=row, column=1, value=result.experiment_id)
            ws.cell(row=row, column=2, value=result.method_name)
            ws.cell(row=row, column=3, value=result.noise_type)
            ws.cell(row=row, column=4, value=result.inference_time_ms)
            ws.cell(row=row, column=5, value=result.memory_mb)
            ws.cell(row=row, column=6, value=result.overall_score)
            ws.cell(row=row, column=7, value=result.timestamp)
            row += 1
        
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[chr(64 + col)].width = 18
        
        ws.auto_filter.ref = ws.dimensions
        
        line_chart = LineChart()
        line_chart.title = "Inference Time Comparison"
        line_chart.x_axis.title = "Method"
        line_chart.y_axis.title = "Time (ms)"
        line_chart.style = 12
        
        data_range = Reference(ws, min_col=4, min_row=1, max_row=row - 1)
        cats = Reference(ws, min_col=2, min_row=2, max_row=row - 1)
        
        line_chart.add_data(data_range, titles_from_data=True)
        line_chart.set_categories(cats)
        line_chart.dataLabels = DataLabelList()
        line_chart.dataLabels.showVal = True
        
        ws.add_chart(line_chart, "I2")
    
    def _add_overall_summary_sheet(self, wb: Workbook, results: List[ComparisonResult]):
        ws = wb.create_sheet(title="Overall Summary")
        
        header_style = self._create_header_style()
        
        headers = ['Experiment ID', 'Method', 'Noise Type',
                   'Position RMSE', 'Angle RMSE', 'Emissivity RMSE',
                   'Classification Acc (%)', 'F1 Score',
                   'Inference Time (ms)', 'Overall Score']
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            for key, value in header_style.items():
                setattr(cell, key, value)
        
        row = 2
        for result in results:
            ws.cell(row=row, column=1, value=result.experiment_id)
            ws.cell(row=row, column=2, value=result.method_name)
            ws.cell(row=row, column=3, value=result.noise_type)
            ws.cell(row=row, column=4, value=result.rmse_distance)
            ws.cell(row=row, column=5, value=result.rmse_angle)
            ws.cell(row=row, column=6, value=result.rmse_emissivity)
            ws.cell(row=row, column=7, value=result.classification_acc)
            ws.cell(row=row, column=8, value=result.f1_score)
            ws.cell(row=row, column=9, value=result.inference_time_ms)
            ws.cell(row=row, column=10, value=result.overall_score)
            row += 1
        
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[chr(64 + col)].width = 16
        
        ws.auto_filter.ref = ws.dimensions
        
        bar_chart = BarChart()
        bar_chart.title = "Overall Score Comparison"
        bar_chart.x_axis.title = "Method"
        bar_chart.y_axis.title = "Score"
        bar_chart.style = 10
        bar_chart.y_axis.scaling.min = 0
        bar_chart.y_axis.scaling.max = 1
        
        data_range = Reference(ws, min_col=10, min_row=1, max_row=row - 1)
        cats = Reference(ws, min_col=2, min_row=2, max_row=row - 1)
        
        bar_chart.add_data(data_range, titles_from_data=True)
        bar_chart.set_categories(cats)
        bar_chart.dataLabels = DataLabelList()
        bar_chart.dataLabels.showVal = True
        
        ws.add_chart(bar_chart, "L2")
    
    def export_dataframe(self, df: pd.DataFrame, sheet_name: str = "Data",
                         filename: str = None, add_charts: bool = True) -> str:
        if filename is None:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"data_export_{timestamp}.xlsx"
        
        filepath = os.path.join(self.output_dir, filename)
        
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name
        
        header_style = self._create_header_style()
        
        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                if r_idx == 1:
                    for key, value in header_style.items():
                        setattr(cell, key, value)
        
        for col in range(1, len(df.columns) + 1):
            ws.column_dimensions[chr(64 + col)].width = 15
        
        ws.auto_filter.ref = ws.dimensions
        
        if add_charts and len(df) > 1:
            numeric_cols = df.select_dtypes(include=[np.number]).columns
            if len(numeric_cols) >= 2:
                bar_chart = BarChart()
                bar_chart.title = f"{sheet_name} Data Summary"
                bar_chart.style = 10
                
                data_range = Reference(ws, min_col=2, min_row=1, max_row=len(df) + 1)
                cats = Reference(ws, min_col=1, min_row=2, max_row=len(df) + 1)
                
                bar_chart.add_data(data_range, titles_from_data=True)
                bar_chart.set_categories(cats)
                
                ws.add_chart(bar_chart, f"{chr(65 + len(df.columns) + 1)}2")
        
        wb.save(filepath)
        
        return filepath

def create_excel_exporter(output_dir: str = "./experiment_results") -> ExcelExporter:
    return ExcelExporter(output_dir=output_dir)
